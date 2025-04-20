import random
import time
import httpx
import json
from openpyxl import Workbook
from datetime import datetime


class Kream:
    def __init__(self):
        self.header = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/134.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
            'Accept-Language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
            'Accept-Encoding': 'gzip, deflate, br, zstd',
            'Origin': 'https://kream.co.kr/',
            'Sec-Ch-Ua-Platform': '"Windows"',
            'Connection': 'keep-alive',
            'Sec-Fetch-Dest': 'empty',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Site': 'same-site',
            'Priority': 'u=0, i'
        }

    def cookie_make(self, headers):
        cookie = []
        headers = str(headers).replace(": '", "").split("'set-cookie'")[1:]
        for i in headers:
            cookie.append(i.split("; ")[0].replace(", '", ""))
        if len(cookie) == 0:
            cookie = ""
        elif len(cookie) != 1:
            cookie = "; ".join(cookie)
        else:
            cookie = cookie[0]
        return cookie

    def login(self, htx, data):
        self.header["X-KREAM-API-VERSION"] = htx.split("apiVersion:")[-1].split(",")[0]
        self.header["X-KREAM-WEB-REQUEST-SECRET"] = htx.split('webRequestSecret:"')[-1].split('"')[0]
        self.header["X-KREAM-WEB-BUILD-VERSION"] = htx.split('sentry-release=')[-1].split(',')[0]
        self.header["X-KREAM-CLIENT-DATETIME"] = f'{datetime.now().strftime("%Y%m%d%H%M%S%z")}+0900'
        self.header["X-KREAM-DEVICE-ID"] = "web;53a1f669-a668-4925-8524-c7cc9120486b"
        self.header["Content-Type"] = "application/json"
        self.header["Accept"] = "application/json, text/plain, */*"
        login_data = {
              "email": data["id"],
              "password": data["pw"]
            }

        htx = httpx.post("https://api.kream.co.kr/api/auth/login", headers=self.header, data=json.dumps(login_data))
        self.header["Cookie"] = f'{self.header["Cookie"]}; {self.cookie_make(htx.headers)}'
        self.header["Authorization"] = f'Bearer {htx.json()["access_token"]}'
        self.header["Accept"] = "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7"
        self.header["Content-Type"] = "text/html; charset=utf-8"
        self.header["Sec-Fetch-Dest"] = 'document'
        self.header["Sec-Fetch-Mode"] = 'navigate'
        self.header["Sec-Fetch-Site"] = 'same-origin'
        self.header["Accept-Encoding"] = 'gzip, deflate, br, zstd'
        self.header["Accept-Language"] = 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7'

    def get_sell_list(self):
        page = 1
        num = 1
        write_wb = Workbook()
        write_ws = write_wb.active
        while True:
            htx = httpx.get(f"https://api.kream.co.kr/api/o/asks/?cursor={page}&tab=finished&status=all", headers=self.header)
            for i in htx.json()["items"]:
                if i["display_type"] == "product_list_info_action":
                    product_url = f'https://api.kream.co.kr/api/m/asks/{i["actions"][0]["value"].split("/")[-1]}'
                    try:
                        htxx = httpx.get(product_url, headers=self.header).json()
                        if htxx["product"]["release"]["local_price"] is None:
                            price = htxx["product"]["release"]["original_price"]
                        else:
                            price = htxx["product"]["release"]["local_price"]
                    except:
                        htxx = httpx.get(product_url, headers=self.header).json()
                        if htxx["product"]["release"]["local_price"] is None:
                            price = htxx["product"]["release"]["original_price"]
                        else:
                            price = htxx["product"]["release"]["local_price"]
                    write_ws.cell(num + 2, 1, num)
                    write_ws.cell(num + 2, 2, htxx["product"]["release"]["name"])
                    write_ws.cell(num + 2, 3, htxx["product_option"]["name"])
                    write_ws.cell(num + 2, 4, htxx["product"]["release"]["style_code"])
                    write_ws.cell(num + 2, 5, int(htxx["price_breakdown"]["price"]))
                    write_ws.cell(num + 2, 6, int(htxx["price_breakdown"]["processing_fee"]["value"]))
                    write_ws.cell(num + 2, 7, int(htxx["price_breakdown"]["total_payout"]))
                    write_ws.cell(num + 2, 8, price)
                    write_ws.cell(num + 2, 9, htxx["date_paid"])
                    write_ws.cell(num + 2, 10, htxx["status_display"])
                    write_ws.cell(num + 2, 11, f'https://kream.co.kr/my/selling/{i["actions"][0]["value"].split("/")[-1]}')
                    write_wb.save("File/Data.xlsx")
                    print(f'{htxx["product"]["release"]["name"]}|{htxx["product"]["release"]["style_code"]}|{htxx["product_option"]["name"]}|{int(htxx["price_breakdown"]["price"])}|{int(htxx["price_breakdown"]["processing_fee"]["value"])}|{int(htxx["price_breakdown"]["total_payout"])}|{htxx["date_paid"]}')
                    time.sleep(random.uniform(0, 3))
                    num += 1
            if htx.json()["next_cursor"] is None:
                break
            page += 1

    def run(self, data):
        htx = httpx.get("https://kream.co.kr/login", headers=self.header, verify=False)
        self.header["Cookie"] = self.cookie_make(htx.headers)
        self.login(htx.text, data)
        self.get_sell_list()


data = {
    "id": "크림 계정",
    "pw": "크림 암호"
}
Kream().run(data)
